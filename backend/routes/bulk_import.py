"""
Bulk Import Routes for Orders
Handles Excel/CSV file uploads and mass order creation
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timezone
from typing import List, Dict, Any
import os
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import secrets

from models import User, Order, AddressInfo, Organization, OrderStatus, PaymentStatus
from auth_utils import verify_token

logger = logging.getLogger(__name__)
router = APIRouter()

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
db_name = os.environ.get('DB_NAME', 'beyond_express_db')
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

# Auth dependency
from fastapi import Request, Cookie
from typing import Optional

async def get_current_user(request: Request, session_token: Optional[str] = Cookie(None)) -> User:
    """Auth dependency"""
    token = session_token
    
    if not token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
    
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session_doc = await db.sessions.find_one({"session_token": token}, {"_id": 0})
    if session_doc:
        if datetime.fromisoformat(session_doc['expires_at']) > datetime.now(timezone.utc):
            user_doc = await db.users.find_one({"id": session_doc['user_id']}, {"_id": 0})
            if user_doc:
                return User(**user_doc)
    
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    
    return User(**user_doc)

@router.get("/template")
async def download_template(current_user: User = Depends(get_current_user)):
    """
    Generate and download Excel template for bulk import
    """
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commandes"
        
        # Define columns
        headers = [
            "Nom Client",
            "Téléphone",
            "Wilaya",
            "Commune",
            "Adresse",
            "Produit (Description)",
            "Prix COD (DZD)",
            "Remarque (Optionnel)"
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add example row
        example_row = [
            "Ahmed Bensalah",
            "0555123456",
            "Alger",
            "Bab Ezzouar",
            "Cité 123, Rue de la Liberté",
            "Téléphone Samsung Galaxy A54",
            "35000",
            "Appeler avant livraison"
        ]
        
        for col_num, value in enumerate(example_row, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as download
        headers_response = {
            'Content-Disposition': 'attachment; filename="template_import_commandes.xlsx"'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response
        )
    
    except Exception as e:
        logger.error(f"Error generating template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/bulk-import")
async def bulk_import_orders(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Import orders from Excel/CSV file
    Auto-calculates shipping cost and net to merchant
    """
    try:
        # Check file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            raise HTTPException(status_code=400, detail="Format de fichier non supporté. Utilisez .xlsx, .xls ou .csv")
        
        # Read file content
        contents = await file.read()
        
        # Parse file
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(contents))
        else:
            df = pd.read_excel(BytesIO(contents))
        
        # Validate columns
        required_columns = ["Nom Client", "Téléphone", "Wilaya", "Commune", "Adresse", "Produit (Description)", "Prix COD (DZD)"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Colonnes manquantes: {', '.join(missing_columns)}"
            )
        
        # Get organization info for sender
        org = await db.organizations.find_one({}, {"_id": 0})
        if not org:
            org = Organization().model_dump()
            org['created_at'] = org['created_at'].isoformat()
            await db.organizations.insert_one(org)
        
        sender_info = AddressInfo(
            name=org['name'],
            phone=org['phone'],
            address=org['address'],
            wilaya="Batna",
            commune="Batna"
        )
        
        # Get all pricing data
        pricing_data = await db.pricing_table.find({}, {"_id": 0}).to_list(1000)
        pricing_map = {}
        for p in pricing_data:
            key = f"{p['wilaya'].lower().strip()}_{p['delivery_type']}"
            pricing_map[key] = p['price']
        
        # Process orders
        results = {
            "success": [],
            "errors": [],
            "total": len(df),
            "created": 0,
            "failed": 0
        }
        
        now = datetime.now(timezone.utc)
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row["Nom Client"]) or pd.isna(row["Téléphone"]):
                    continue
                
                # Extract data
                client_name = str(row["Nom Client"]).strip()
                phone = str(row["Téléphone"]).strip()
                wilaya = str(row["Wilaya"]).strip()
                commune = str(row["Commune"]).strip()
                address = str(row["Adresse"]).strip()
                product = str(row["Produit (Description)"]).strip()
                cod_amount = float(row["Prix COD (DZD)"])
                notes = str(row.get("Remarque (Optionnel)", "")).strip() if "Remarque (Optionnel)" in row else ""
                
                # Calculate shipping cost from pricing table
                shipping_cost = 0.0
                delivery_type = "home"  # Default to home delivery
                
                # Try to find pricing
                pricing_key = f"{wilaya.lower()}_{delivery_type}"
                if pricing_key in pricing_map:
                    shipping_cost = pricing_map[pricing_key]
                else:
                    # Try without accents/special chars (basic normalization)
                    results["errors"].append({
                        "row": index + 2,  # +2 because Excel rows start at 1 and we have header
                        "error": f"Prix de livraison non trouvé pour {wilaya}. Utilisation de 0 DZD.",
                        "data": {"client": client_name, "wilaya": wilaya}
                    })
                
                # Calculate net to merchant
                net_to_merchant = cod_amount - shipping_cost
                
                # Generate tracking ID and PIN
                tracking_id = f"BEX-{secrets.token_hex(6).upper()}"
                import random
                pin_code = str(random.randint(1000, 9999))
                
                # Create recipient address
                recipient = AddressInfo(
                    name=client_name,
                    phone=phone,
                    address=address,
                    wilaya=wilaya,
                    commune=commune
                )
                
                # Create order
                order = Order(
                    tracking_id=tracking_id,
                    user_id=current_user.id,
                    recipient=recipient,
                    sender=sender_info,
                    cod_amount=cod_amount,
                    description=product,
                    service_type="E-COMMERCE",
                    delivery_type="Livraison à Domicile",
                    status=OrderStatus.IN_STOCK,
                    pin_code=pin_code,
                    shipping_cost=shipping_cost,
                    net_to_merchant=net_to_merchant,
                    payment_status=PaymentStatus.UNPAID,
                    created_at=now,
                    updated_at=now
                )
                
                # Save to database
                order_dict = order.model_dump()
                order_dict['created_at'] = order_dict['created_at'].isoformat()
                order_dict['updated_at'] = order_dict['updated_at'].isoformat()
                if order_dict.get('collected_date'):
                    order_dict['collected_date'] = order_dict['collected_date'].isoformat()
                if order_dict.get('transferred_date'):
                    order_dict['transferred_date'] = order_dict['transferred_date'].isoformat()
                
                await db.orders.insert_one(order_dict)
                
                results["success"].append({
                    "row": index + 2,
                    "tracking_id": tracking_id,
                    "client": client_name,
                    "cod": cod_amount,
                    "shipping": shipping_cost,
                    "net": net_to_merchant
                })
                results["created"] += 1
                
            except Exception as e:
                results["errors"].append({
                    "row": index + 2,
                    "error": str(e),
                    "data": row.to_dict()
                })
                results["failed"] += 1
                logger.error(f"Error processing row {index + 2}: {str(e)}")
        
        logger.info(f"✅ Bulk import completed: {results['created']} created, {results['failed']} failed")
        
        return results
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in bulk import: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
